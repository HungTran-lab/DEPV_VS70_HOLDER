import os
import sys
import csv
import serial
import serial.tools.list_ports
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QComboBox, QLineEdit, QDesktopWidget, QMessageBox, QTableWidgetItem
from PyQt5.uic import loadUi
from PyQt5.QtGui import QPixmap, QImage, QPainter, QColor, QFont, QIcon
from PyQt5.QtCore import QTimer, QDateTime, Qt, pyqtSignal
from PyQt5.QtPrintSupport import QPrinter
from datetime import datetime
from PyQt5.QtGui import QTextCursor

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

from PyQt5.QtWidgets import QMessageBox
import qrcode
import random
#from read_from_com import *

class MyWindow(QMainWindow):
    # Tạo signal để gửi ngày, tháng, năm
    date_signal = pyqtSignal(str, str, str) 

    def __init__(self):
        super().__init__()

        # Load UI từ file .ui
        loadUi("gui2.ui", self)

        self.load_config()

        # Thay đổi tiêu đề cửa sổ
        self.setWindowTitle("FT Assy Charger Base")


        # Đặt icon cho cửa sổ
        icon_path = 'stick.png'  # Đảm bảo đường dẫn icon đúng
        icon_pixmap = QPixmap(icon_path).scaled(90, 100)
        self.setWindowIcon(QIcon(icon_pixmap))

        self.setFixedSize(1380, 670)
        self.center_window()


        #  add Table---------------------------
        self.data_table.setColumnCount(4)  # cột
        self.data_table.setHorizontalHeaderLabels(["No.", "Time", "ADC Value", "Result"])

        # Danh sách lưu dữ liệu (tối đa 5 phần tử)
        self.data_list = [4]

         # Kết nối tín hiệu (signal) với hành động
        self.actionManual.triggered.connect(self.show_manual_message)
        self.actionVer.triggered.connect(self.show_about_message)
        self.actionInfor.triggered.connect(self.show_infor_message)

        # Khởi tạo các widget (nếu chưa được set trong .ui)
        self.connect_button = self.findChild(QPushButton, 'connect_button')
        self.comboBox_com_ports = self.findChild(QComboBox, 'comboBox_com_ports')
        self.receive_data = self.findChild(QLineEdit, 'receive_data')
        # Các widget hiển thị trạng thái, thời gian, QR code,... được giả định đã có trong file gui.ui
        # Ví dụ: self.label_time, self.label_date, self.nam, self.thang, self.ngay,
        # self.display (QPlainTextEdit), self.ok_count (QLabel), self.ng_count (QLabel),
        # self.total_count (QLabel), self.value_adc (QLabel), self.value (QLabel),
        # self.label_qr_code (QLabel), self.comboBox (QComboBox) cho additional_text,...

        # Khởi tạo biến đếm
        self.ok_count_value = 0000
        self.ng_count_value = 0000
        self.total_count_value = 0000
        self.value_adc1 = 0

        self.qr_print.setReadOnly(True)
        self.dept.setReadOnly(True)
        self.company.setReadOnly(True)
        self.status.setReadOnly(True)

        # Nút Connect khi nhấn vào sẽ gọi hàm connect_com
        self.connect_button.clicked.connect(self.connect_com)

        # Các sự kiện cho nút
        self.make_qr.clicked.connect(self.make_qr_code1)
        self.print_qr.clicked.connect(self.print_qr_code)
        #self.save_data.clicked.connect(self.save_qlineedit_to_excel)

        # Biến để lưu ảnh QR code
        self.qr_image = None

        # Thiết lập timer để cập nhật thời gian và ngày
        timer = QTimer(self)
        timer.timeout.connect(self.update_time)
        timer.start(1000)  # Cập nhật mỗi 1 giây

        self.serial_timer = QTimer(self)
        self.serial_timer.timeout.connect(self.read_from_com)
        self.serial_timer.start(100)  # Kiểm tra mỗi 100ms

        # Lấy danh sách các cổng COM khả dụng
        self.populate_com_ports()

        # Kết nối sự kiện comboBox để cập nhật cổng COM
        self.comboBox_com_ports.currentIndexChanged.connect(self.update_serial_port)

        # Khởi tạo kết nối serial
        self.serial_connection = None

        # <<<< THÊM: Tải giá trị counter từ file data.csv và cập nhật lên giao diện
        self.load_counter()
        self.ok_count.setText(f"{self.ok_count_value:04d}")  # show value of OK to GUI
        self.ng_count.setText(f"{self.ng_count_value:04d}")  # show value of NG to GUI
        self.total_count.setText(f"{self.total_count_value:04d}")


        self.last_state = "NONE"   # lưu trạng thái trước đó

    def update_time(self):
        current_datetime = QDateTime.currentDateTime()
        time_str = current_datetime.toString("HH:mm:ss")
        self.label_time.setText(time_str)
        date_str = current_datetime.toString("dd-MM-yyyy")
        self.label_date.setText(date_str)

        # Cập nhật năm
        year = current_datetime.toString("yyyy")
        year_mapping = {
            "2025": "Y",
            "2026": "L",
            "2027": "P",
        }
        self.year_display = year_mapping.get(year, year)
        self.nam.setText(self.year_display)

        # Cập nhật tháng
        month = current_datetime.toString("M")
        month_mapping = {
            "10": "A",
            "11": "B",
            "12": "C"
        }
        self.month_display = month_mapping.get(month, month)
        self.thang.setText(self.month_display)

        # Cập nhật ngày
        day = current_datetime.toString("d")
        day_mapping = {
            "1": "1",
            "2": "2",
            "3": "3",
            "4": "4",
            "5": "5",
            "6": "6",
            "7": "7",
            "8": "8",
            "9": "9",
            "10": "A",
            "11": "B",
            "12": "C",
            "13": "D",
            "14": "E",
            "15": "F",
            "16": "G",
            "17": "H",
            "18": "J",
            "19": "K",
            "20": "L",
            "21": "M",
            "22": "N",
            "23": "P",
            "24": "R",
            "25": "S",
            "26": "T",
            "27": "V",
            "28": "W",
            "29": "X",
            "30": "Y",
            "31": "Z",
        }
        self.day_display = day_mapping.get(day, day)
        self.ngay.setText(self.day_display)

    def load_config(self):
        try:
            with open('config.csv', mode='r') as file:
                csv_reader = csv.reader(file)
                for row in csv_reader:
                    key, value = row
                    # if key == "id":
                        # self.id.setText(value)
                    if key == "name":
                        self.name.setText(value)
                    elif key == "vendor code":
                        self.dept.setText(value)
                    elif key == "part code":
                        self.company.setText(value)
        except FileNotFoundError:
            print("File CSV không tồn tại.")
        except Exception as e:
            print(f"Lỗi khi đọc file CSV: {e}")
    def load_counter(self):
        """
        Đọc giá trị counter từ file data.csv.
        Nếu file tồn tại và được cập nhật trong ngày, dùng giá trị trong file.
        Nếu file không tồn tại hoặc đã cũ (khác ngày), reset counter về 0 và lưu lại.
        """
        if os.path.exists("data.csv"):
            try:
                mod_time = os.path.getmtime("data.csv")
                mod_date = datetime.fromtimestamp(mod_time).date()
                today = datetime.now().date()
                if mod_date == today:
                    with open("data.csv", "r", newline='') as f:
                        reader = csv.reader(f)
                        for row in reader:
                            if row[0] == "OK":
                                self.ok_count_value = int(row[1])
                            elif row[0] == "NG":
                                self.ng_count_value = int(row[1])
                            elif row[0] == "Total":
                                self.total_count_value = int(row[1])
                               
                else:
                    # Nếu file không được cập nhật trong ngày, reset counter
                    self.ok_count_value =  0000  # just change from 0
                    self.ng_count_value = 0 # rst 
                    self.save_counter()
            except Exception as e:
                print("Error loading counter:", e)
                self.ok_count_value =  0000  # just change from 0
                self.ng_count_value=0

        else:
            self.ok_count_value =  0000  # just change from 0
            self.ng_count_value =0
            self.save_counter()
    def connect_com(self):
        selected_port = self.comboBox_com_ports.currentText()
        if selected_port:
            if self.serial_connection and self.serial_connection.is_open:
                self.serial_connection.close()
                self.status.setText(f"Disconnected from {selected_port}")
                self.status.setStyleSheet("color: orange;")
            else:
                try:
                    # dùng baudrate 115200
                    self.serial_connection = serial.Serial(selected_port, 115200, timeout=1)
                    self.status.setText(f"Connected to {selected_port} - 115200")
                    self.status.setStyleSheet("color: green;")
                except serial.SerialException as e:
                    self.status.setText(f"Failed to connect: {e}")
                    self.status.setStyleSheet("color: red;")


    def populate_com_ports(self):
        ports = serial.tools.list_ports.comports()
        self.comboBox_com_ports.clear()
        for port in ports:
            self.comboBox_com_ports.addItem(port.device)

    def update_serial_port(self):
        selected_port = self.comboBox_com_ports.currentText()
        if self.serial_connection and self.serial_connection.is_open:
            self.serial_connection.close()
            print("Disconnected from old port")
            self.status.setText("Disconnected from old port")
            self.status.setStyleSheet("color: orange;")
        if selected_port:
            try:
                self.serial_connection = serial.Serial(selected_port, 9600, timeout=1)
                print(f"Connected to {selected_port}")
                self.status.setText(f"Connected to {selected_port}")
                self.status.setStyleSheet("color: green;")
            except serial.SerialException as e:
                print(f"Failed to connect to {selected_port}: {e}")
                self.status.setText(f"Failed to connect: {e}")
                self.status.setStyleSheet("color: red;")

    
    
    
    # def read_from_com(self):
    def read_from_com(self):
        if not self.serial_connection or not self.serial_connection.is_open:
            return

        try:
            while self.serial_connection.in_waiting > 0:
                line = self.serial_connection.readline().decode('utf-8', errors='ignore').strip()

                if line:
                    # Hiển thị log
                    self.display.appendPlainText(line)
                    self.display.moveCursor(QTextCursor.End)

                    # --- Giới hạn chỉ giữ 10 dòng log ---
                    max_lines = 20
                    text = self.display.toPlainText().splitlines()
                    if len(text) > max_lines:
                        # Xóa các dòng cũ, chỉ giữ 10 dòng cuối
                        text = text[-max_lines:]
                        self.display.setPlainText("\n".join(text))
                        self.display.moveCursor(QTextCursor.End)

                    if line.startswith("Waiting"):
                        self.value.setText("Wait")
                        self.value.setStyleSheet("color: orange; background-color: lightyellow;")
                        self.reset_sensors()   # quay lại màu xanh dương mặc định

                    elif line.startswith("OK"):
                        self.value.setText("OK")
                        self.value.setStyleSheet("color: green; background-color: lightblue;")
                        self.reset_sensors()   # OK → cũng reset về xanh dương

                    elif line.startswith("NG"):
                        self.value.setText("NG")
                        self.value.setStyleSheet("color: red; background-color: lightblue;")

                        # reset về mặc định trước
                        self.reset_sensors()

                        if "data=" in line:
                            data_str = line.split("data=")[1]
                            values = data_str.split(",")

                            for i, val in enumerate(values, start=1):
                                label = getattr(self, f"sc_{i}")
                                if val.strip() == "1":
                                    label.setStyleSheet("background-color: red; border-radius: 20px;")

        except Exception as e:
            print("Serial error:", e)
            self.status.setText("Lỗi cổng COM")
            self.status.setStyleSheet("color: red;")
            self.reconnect_com()

  

    def make_qr_code(self):
         # Lấy các giá trị đã chuyển đổi từ hàm update_time
        year = self.year_display
        month = self.month_display
        day = self.day_display 

        # data1 = self.id.text()
        # data2 = self.name.text()
        data3 = self.dept.text()
        data4 = self.company.text()
        data5 = self.value.text()

        additional_text = self.comboBox.currentText()  # Hàng thứ 2, Part Code

        current_time = datetime.now().strftime("%H:%M:%S")
        current_date = datetime.now().strftime("%d-%m-%Y")

        # <<<< CHÚ Ý: Phần dữ liệu 19 ký tự bên trái đã có, sau đó thêm 4 ký tự số đếm
        counter_str = f"{self.ok_count_value:04d}"
        # Ví dụ: 18 + additional_text + data3 + year + month + day => 19 ký tự, rồi nối counter_str
        qr_data = "".join(["18", additional_text, data3, year, month, day, counter_str])
        print("QR data:", qr_data)  # In ra để kiểm tra

        qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=12, border=2)   #10
        qr.add_data(qr_data)
        qr.make(fit=True)
        qr_pil_image = qr.make_image(fill_color="black", back_color="white")
        
        qr_pil_image = qr_pil_image.resize((236, 236))#236*236
        qr_qimage = QImage(qr_pil_image.size[0], qr_pil_image.size[1], QImage.Format_ARGB32)
        for x in range(qr_pil_image.size[0]):
            for y in range(qr_pil_image.size[1]):
                color = qr_pil_image.getpixel((x, y))
                qr_qimage.setPixelColor(x, y, QColor(color, color, color))

        combined_pixmap = QPixmap(236+20, 295)  #236*295
        combined_pixmap.fill(Qt.white)
        painter = QPainter(combined_pixmap)
        painter.drawImage(40, 0, qr_qimage)   ## X = -10, Y = 0 (dịch trái 10 pixel)
        painter.setFont(QFont("SamsungSharpSans-Bold", 18))
        painter.setPen(Qt.black)
        text_width = painter.fontMetrics().horizontalAdvance(additional_text)
        text_x = (236+20 - text_width) // 2+40    # Căn giữa và dịch trái 10 pixel
        painter.drawText(text_x, 270, additional_text)
        painter.end()

        combined_pixmap.save("qr_code.png")
        self.qr_image = combined_pixmap

        self.label_qr_code.setPixmap(combined_pixmap.scaled(self.label_qr_code.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)) 
        
    def make_qr_code1(self):
        # Lấy các giá trị đã chuyển đổi từ hàm update_time
        year = self.year_display
        month = self.month_display
        day = self.day_display 

        # data1 = self.id.text()
        # data2 = self.name.text()
        data3 = self.dept.text()
        data4 = self.company.text()
        data5 = self.value.text()

        additional_text = self.comboBox.currentText()  # Hàng thứ 2, Part Code (ví dụ: DJ96xxxxxx)
        additional_text_formatted = f"{additional_text[:4]}-{additional_text[4:]}"

        current_time = datetime.now().strftime("%H:%M:%S")
        current_date = datetime.now().strftime("%d-%m-%Y")

        # <<<< CHÚ Ý: Thêm dấu '-' sau 4 ký tự đầu (DJ96-xxxxxx)
        # Ví dụ: additional_text = "DJ96xxxxxx" → "DJ96-xxxxxx"
        
        
        
            

        counter_str = f"{self.ok_count_value:04d}"
        qr_data = "".join(["18", additional_text, data3, year, month, day, counter_str])  # Sử dụng additional_text_formatted
        #print("QR data:", qr_data)  # In ra để kiểm tra
        self.qr_print.setText(qr_data)
        qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=12, border=2)
        qr.add_data(qr_data)
        qr.make(fit=True)
        qr_pil_image = qr.make_image(fill_color="black", back_color="white")
        
        qr_pil_image = qr_pil_image.resize((236, 236))
        qr_qimage = QImage(qr_pil_image.size[0], qr_pil_image.size[1], QImage.Format_ARGB32)
        for x in range(qr_pil_image.size[0]):
            for y in range(qr_pil_image.size[1]):
                color = qr_pil_image.getpixel((x, y))
                qr_qimage.setPixelColor(x, y, QColor(color, color, color))

        combined_pixmap = QPixmap(236+20, 295)    #295 --? 280
        combined_pixmap.fill(Qt.white)
        painter = QPainter(combined_pixmap)
        painter.drawImage(10, 0, qr_qimage)
        painter.setFont(QFont("SamsungSharpSans-Bold", 18))
        painter.setPen(Qt.black)
        text_width = painter.fontMetrics().horizontalAdvance(additional_text_formatted)  # Sử dụng additional_text_formatted
        text_x = (236 - text_width) // 2 + 10
        painter.drawText(text_x, 255, additional_text_formatted)  # Hiển thị định dạng mới  270--> 250
        painter.end()

        combined_pixmap.save("qr_code.png")
        self.qr_image = combined_pixmap
        self.label_qr_code.setPixmap(combined_pixmap.scaled(self.label_qr_code.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation))
   
   
    #-------------------Export to .CSV File--------------------------------------------
    def save_qlineedit_to_csv(self):
        read_adc = self.value_adc.text()  # Giá trị ADC
        status = self.value.text()        # Trạng thái OK/NG

        #if not read_adc or not status:
        #    QMessageBox.warning(self, "Cảnh báo", "Không có dữ liệu nào để lưu!")
        #    return

        # Tạo thư mục nếu chưa có
        os.makedirs("data", exist_ok=True)

        # Đặt tên file theo ngày
        save_path = os.path.join("data", f"adc_data_{datetime.now().strftime('%Y-%m-%d')}.csv")

        file_exists = os.path.exists(save_path)

        try:
            with open(save_path, mode='a', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)

                # Nếu file mới tạo, ghi dòng tiêu đề
                if not file_exists:
                    writer.writerow(["No.", "ADC Value", "Time and Date", "Status"])

                # Đếm số dòng hiện tại để xác định STT
                with open(save_path, mode='r', encoding='utf-8') as count_file:
                    row_count = sum(1 for row in count_file) - 1  # Trừ 1 vì có dòng tiêu đề

                times = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                writer.writerow([row_count + 1, read_adc, times, status])

            # QMessageBox.information(self, "Thành công", f"Dữ liệu đã được lưu tại {save_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể lưu dữ liệu: {str(e)}")
    
    def save_qlineedit_to_excel(self):
        read_adc = self.value_adc.text()  # Giá trị ADC
        status = self.value.text()  # Trạng thái OK/NG

        # if not read_adc or not status:
        #      QMessageBox.warning(self, "Cảnh báo", "Không có dữ liệu nào để lưu!")
        #      return
    
        # Đặt tên file theo ngày
        #save_path = f"adc_data_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        save_path = os.path.join("data", f"adc_data_{datetime.now().strftime('%Y-%m-%d')}.xlsx")


        try:
              # Kiểm tra xem file đã tồn tại hay chưa
            if os.path.exists(save_path):
                wb = load_workbook(save_path)  # Mở file hiện có
                ws = wb.active
            else:
                wb = Workbook()  # Tạo mới nếu chưa tồn tại
                ws = wb.active
                ws.title = "Adc_data"
                ws.append(["No.", "ADC Value", "Time and Date", "Status"])  # Ghi tiêu đề
        
            next_row = ws.max_row + 1
            times = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            ws.append([next_row - 1, read_adc, times, status])  # `next_row - 1` để STT bắt đầu từ 1

            wb.save(save_path)  # Lưu file Excel
           #QMessageBox.information(self, "Thành công", f"Dữ liệu đã được lưu tại {save_path}")

        except Exception as e:
         QMessageBox.critical(self, "Lỗi", f"Không thể lưu dữ liệu: {str(e)}")

    def print_qr_code(self):
        # Kiểm tra trạng thái radio button
        if self.enable_print.isChecked():  # Nếu được chọn → KHÔNG cho in
             #print("Printing disabled: Radio button is checked")
             #QMessageBox.information(self, "Thông báo", "Chức năng in đã bị tắt")
             return  # Thoát hàm, không in
        
        if self.qr_image is not None:
            printer = QPrinter(QPrinter.HighResolution)
            printer.setOutputFormat(QPrinter.NativeFormat)
            painter = QPainter(printer)
            
            rect = painter.viewport()
            size = self.qr_image.size()
            size.scale(rect.size(), Qt.KeepAspectRatio)
            painter.setViewport(rect.x(), rect.y(), size.width(), size.height())
            painter.setWindow(self.qr_image.rect())

            painter.drawPixmap(0, 0, self.qr_image)
            painter.end()
            #print("QR Code is printed successfully.")
        else:
            print("No QR code to print.")

    def add_new_data(self):
        adc_value = self.value_adc.text()
        try:
            adc_value = float(adc_value)
        except ValueError:
            print("Lỗi: Giá trị ADC không hợp lệ!")
            return

        result1 = self.value.text()
        timestamp = QDateTime.currentDateTime().toString("hh:mm:ss")
        
        self.data_table.insertRow(0)
        self.data_table.setItem(0, 0, QTableWidgetItem(str(1)))
        self.data_table.setItem(0, 1, QTableWidgetItem(timestamp))
        self.data_table.setItem(0, 2, QTableWidgetItem(str(adc_value)))
        self.data_table.setItem(0, 3, QTableWidgetItem(str(result1)))

        for i in range(self.data_table.rowCount()):
            self.data_table.setItem(i, 0, QTableWidgetItem(str(i + 1)))

        while self.data_table.rowCount() > 5:
            self.data_table.removeRow(5)

      
    def reconnect_com(self):
     if self.serial_connection:
        self.serial_connection.close()

     selected_port = self.comboBox_com_ports.currentText()
     if selected_port:
        try:
            self.serial_connection = serial.Serial(selected_port, 115200, timeout=1)
            print(f"Reconnected to {selected_port}")
            self.status.setText(f"Reconnected to {selected_port}")
            self.status.setStyleSheet("color: green;")
        except serial.SerialException as e:
            print(f"Failed to reconnect to {selected_port}: {e}")
            self.status.setText("Failed to reconnect")
            self.status.setStyleSheet("color: red;")

    def save_counter(self):
        """
        Lưu giá trị counter vào file data.csv theo định dạng: OK,xxxx
        với xxxx là số có 4 chữ số (ví dụ: 0001, 0234,...)
        """
        try:
              # Tạo thư mục nếu cần
            #os.makedirs(os.path.dirname("data.csv"), exist_ok=True)

            with open("data.csv", "w", newline='') as f:
                writer = csv.writer(f)
                writer.writerow(["OK", f"{self.ok_count_value:04d}"])  # save dong 1  OK,XXXX
                writer.writerow(["NG", f"{self.ng_count_value:04d}"])  # save dong 2  NG,XXXX
                writer.writerow(["Total", f"{self.total_count_value:04d}"])  # save dong 3  NG,XXXX
        except Exception as e:
            print("Error saving counter:", e)



    def center_window(self):
        # Lấy kích thước màn hình
        screen_geometry = QDesktopWidget().screenGeometry()
        screen_width = screen_geometry.width()
        screen_height = screen_geometry.height()

        # Lấy kích thước cửa sổ
        window_width = self.width()
        window_height = self.height()

        # Tính toán tọa độ để cửa sổ ở giữa
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2

        # Đặt vị trí cửa sổ
        self.move(x, y)   

    def show_about_message(self):
        # Hiển thị thông báo khi nhấn vào About
        QMessageBox.information(self, "About", "Ver1.\n      FT Assy Charger Pipe_VS70\n      Oct-25")   

    def show_manual_message(self):
        # Hiển thị thông báo khi nhấn vào Manual
        QMessageBox.information(self, "Manual", "1111: OK\n"
        "0100: NG_Trên_Phải\n"
        "0111: NG_Trên_Trái\n"
        "0010: NG_Dưới_Phải\n"
        "1110: NG_Dưới_Phải\n"

        "1010: NG_Ngược, Out 2Pin Dưới\n"
        "0101: NG_Ngược, Out 2Pin Trên")
               

    def show_infor_message(self):
        QMessageBox.information(self, "Contact PIC.", "songhung.tr \nVC/RD-Stick Team\nMobi: 03750311**")       
    # ===== THÊM HÀM: load_counter và save_coun


    def reset_sensors(self):
        for i in range(1, 10):   # sc_1 -> sc_9
            label = getattr(self, f"sc_{i}")
            label.setStyleSheet("background-color: rgb(0, 0, 255); border-radius: 20px;")







if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MyWindow()
    window.show()
    sys.exit(app.exec_())
